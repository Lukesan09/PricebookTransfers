from openpyxl import Workbook, load_workbook
import sys


def create_nrs_spreadsheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "pricebook"
    headers = ["Upc", "Department", "qty", "cents", "incltaxes", "inclfees", "Name", "size",
           "ebt", "byweight", "Fee Multiplier", "cost_qty", "cost_cents"]
    ws.append(headers)
    wb.save("nrs.xlsx")


def transfer_upc(clover_ws, nrs_ws):
    upc_col = clover_ws['I']

    for i in range(1, len(upc_col)):
        try:
            nrs_ws['A'+str(i+1)] = '="' + upc_col[i].value + '"'
        except:
            print('nothing')
    

def set_qty(nrs_ws):
    count = len(nrs_ws['A']) - 1 # num of items, subtracting first row
    
    for i in range(2, count+2): # start at row 2
        nrs_ws['C'+str(i)] = 1


def transfer_price(clover_ws, nrs_ws):
    price_col = clover_ws['D'] 

    for i in range(1, len(price_col)):
        try:
            nrs_ws['D'+str(i+1)] = round(price_col[i].value*100)
        except TypeError:
            print("no price")


def taxes_fees_ebt_byweight(nrs_ws):
    cols = ['E', 'F', 'I', 'J']

    count = len(nrs_ws['A']) - 1 # num of items, subtracting first row

    for col in cols:
        for i in range (2, count+2): # start at row 2
            nrs_ws[col+str(i)] = 'n'


def transfer_name(clover_ws, nrs_ws):
    name_col = clover_ws['B']

    for i in range(1, len(name_col)):
        nrs_ws['G'+str(i+1)] = name_col[i].value


def set_fee_multiplier(nrs_ws):
    count = len(nrs_ws['A']) - 1 # num of items, subtracting first row
    
    for i in range(2, count+2): # start at row 2
        nrs_ws['K'+str(i)] = 1


def set_cost_qty(nrs_ws):
    count = len(nrs_ws['A']) - 1 # num of items, subtracting first row
    
    for i in range(2, count+2): # start at row 2
        nrs_ws['L'+str(i)] = 0


def set_cost_cents(nrs_ws):
    count = len(nrs_ws['A']) - 1 # num of items, subtracting first row
    
    for i in range(2, count+2): # start at row 2
        nrs_ws['M'+str(i)] = 0


def set_size(nrs_ws):
    count = len(nrs_ws['A']) - 1 # num of items, subtracting first row
    
    for i in range(2, count+2): # start at row 2
        nrs_ws['H'+str(i)] = '="' + str(0) + '"'


def get_department(clover_ws, nrs_ws):
    nrs_names = nrs_ws['G']
    sheet_values = tuple(clover_ws.rows)
    row_num = 0

    counter = 0

    for name in nrs_names:
        counter += 1
        print(counter)
        row_num += 1
        for col in sheet_values:
            for val in col:
                if name.value == val.value:
                    column = val.column_letter
                    category = clover_ws[column+str(1)].value
                    nrs_ws['B'+str(row_num)] = category
                    continue


def main():
    create_nrs_spreadsheet()
    clover_wb = load_workbook("inventory.xlsx")
    nrs_wb = load_workbook("nrs.xlsx")

    clover_ws = clover_wb.active
    clover_categories = clover_wb["Categories"]
    nrs_ws = nrs_wb.active

    transfer_upc(clover_ws, nrs_ws)
    set_qty(nrs_ws)
    transfer_price(clover_ws, nrs_ws)
    taxes_fees_ebt_byweight(nrs_ws)
    transfer_name(clover_ws, nrs_ws)
    set_fee_multiplier(nrs_ws)
    set_cost_qty(nrs_ws)
    set_cost_cents(nrs_ws)
    set_size(nrs_ws)
    get_department(clover_categories, nrs_ws)

    nrs_wb.save("nrs.xlsx")
    sys.exit()

if __name__ == "__main__":
    main()