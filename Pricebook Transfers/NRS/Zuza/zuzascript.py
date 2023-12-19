from openpyxl import Workbook, load_workbook

count = 0

def create_nrs_spreadsheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "pricebook"
    headers = ["Upc", "Department", "qty", "cents", "incltaxes", "inclfees", "Name", "size",
           "ebt", "byweight", "Fee Multiplier", "cost_qty", "cost_cents"]
    ws.append(headers)
    wb.save("nrs.xlsx")

def set_qty(nrs_ws):
    for i in range(2, count+2): # start at row 2
        nrs_ws['C'+str(i)] = 1


def set_fee_multiplier(nrs_ws):
    for i in range(2, count+2): # start at row 2
        nrs_ws['K'+str(i)] = 1


def set_cost_qty(nrs_ws):
    for i in range(2, count+2): # start at row 2
        nrs_ws['L'+str(i)] = 0


def set_cost_cents(nrs_ws):
    for i in range(2, count+2): # start at row 2
        nrs_ws['M'+str(i)] = 0


def set_size(nrs_ws):
    for i in range(2, count+2): # start at row 2
        nrs_ws['H'+str(i)] = '="' + str(0) + '"'


def taxes_fees_ebt_byweight(nrs_ws):
    cols = ['E', 'F', 'I', 'J']

    for col in cols:
        for i in range (2, count+2): # start at row 2
            nrs_ws[col+str(i)] = 'n'


def transfer_name(zuza_ws, nrs_ws):
    name_col = zuza_ws['A']

    for i in range(1, len(name_col)):
        nrs_ws['G'+str(i+1)] = name_col[i].value


def transfer_price(zuza_ws, nrs_ws):
    price_col = zuza_ws['O']

    for i in range(1, len(price_col)):
        try:
            nrs_ws['D'+str(i+1)] = round(price_col[i].value*100)
        except TypeError:
            print("no price")


def transfer_upc(zuza_ws, nrs_ws):
    upc_col = zuza_ws['D']

    for i in range(1, len(upc_col)):
        nrs_ws['A'+str(i+1)] = upc_col[i].value


def transfer_department(zuza_ws, nrs_ws):
    dep_col = zuza_ws['E']

    for i in range(1, len(dep_col)):
        nrs_ws['B'+str(i+1)] = dep_col[i].value


def main():
    create_nrs_spreadsheet()
    zuza_wb = load_workbook("inventory.xlsm")
    nrs_wb = load_workbook("nrs.xlsx")

    zuza_ws = zuza_wb["Products"]
    nrs_ws = nrs_wb.active

    global count
    count = len(zuza_ws['A']) - 1

    set_qty(nrs_ws)
    taxes_fees_ebt_byweight(nrs_ws)
    set_fee_multiplier(nrs_ws)
    set_cost_qty(nrs_ws)
    set_cost_cents(nrs_ws)
    set_size(nrs_ws)
    transfer_name(zuza_ws, nrs_ws)
    transfer_price(zuza_ws, nrs_ws)
    transfer_upc(zuza_ws, nrs_ws)
    transfer_department(zuza_ws, nrs_ws)

    nrs_wb.save("nrs.xlsx")


if __name__ == "__main__":
    main()
